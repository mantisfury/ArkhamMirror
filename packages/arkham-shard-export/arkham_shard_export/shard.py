"""
Export Shard - Main Shard Implementation

Data export for ArkhamFrame - export documents, entities, claims, and
analysis results in various formats.
"""

import csv
import io
import json
import logging
import os
import tempfile
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional
from uuid import uuid4

import httpx

from arkham_frame import ArkhamShard

from .models import (
    ExportFilter,
    ExportFormat,
    ExportJob,
    ExportOptions,
    ExportResult,
    ExportStatistics,
    ExportStatus,
    ExportTarget,
    FormatInfo,
    TargetInfo,
)

logger = logging.getLogger(__name__)

# Base URL for internal API calls
INTERNAL_API_BASE = "http://127.0.0.1:8100"


class ExportShard(ArkhamShard):
    """
    Export Shard - Export data in various formats.

    This shard provides:
    - Multi-format export (JSON, CSV, PDF, XLSX)
    - Export job management with status tracking
    - Multiple export targets (documents, entities, claims, etc.)
    - File management with expiration
    - Export history and statistics
    """

    name = "export"
    version = "0.1.0"
    description = "Data export in multiple formats (JSON, CSV, PDF, XLSX)"

    def __init__(self):
        super().__init__()  # Auto-loads manifest from shard.yaml
        self.frame = None
        self._db = None
        self._events = None
        self._storage = None
        self._initialized = False
        self._export_dir = None

    async def initialize(self, frame) -> None:
        """Initialize shard with frame services."""
        self.frame = frame
        self._db = frame.database
        self._events = frame.events
        self._storage = getattr(frame, "storage", None)

        # Setup export directory
        self._export_dir = os.path.join(tempfile.gettempdir(), "arkham_exports")
        os.makedirs(self._export_dir, exist_ok=True)

        # Create database schema
        await self._create_schema()

        # Register self in app state for API access
        if hasattr(frame, "app") and frame.app:
            frame.app.state.export_shard = self

        self._initialized = True
        logger.info(f"ExportShard initialized (v{self.version})")

    async def shutdown(self) -> None:
        """Clean shutdown of shard."""
        self._initialized = False
        logger.info("ExportShard shutdown complete")

    def get_routes(self):
        """Return FastAPI router for this shard."""
        from .api import router
        return router

    # === Database Schema ===

    async def _create_schema(self) -> None:
        """Create database tables for export shard."""
        if not self._db:
            logger.warning("Database not available, skipping schema creation")
            return

        await self._db.execute("""
            CREATE TABLE IF NOT EXISTS arkham_export_jobs (
                id TEXT PRIMARY KEY,
                format TEXT NOT NULL,
                target TEXT NOT NULL,
                status TEXT DEFAULT 'pending',

                created_at TEXT,
                started_at TEXT,
                completed_at TEXT,

                file_path TEXT,
                file_size INTEGER,
                download_url TEXT,
                expires_at TEXT,

                error TEXT,
                filters TEXT DEFAULT '{}',
                options TEXT DEFAULT '{}',

                record_count INTEGER DEFAULT 0,
                processing_time_ms REAL DEFAULT 0,
                created_by TEXT DEFAULT 'system',
                metadata TEXT DEFAULT '{}'
            )
        """)

        # Create indexes
        await self._db.execute("""
            CREATE INDEX IF NOT EXISTS idx_export_jobs_status ON arkham_export_jobs(status)
        """)
        await self._db.execute("""
            CREATE INDEX IF NOT EXISTS idx_export_jobs_created ON arkham_export_jobs(created_at)
        """)
        await self._db.execute("""
            CREATE INDEX IF NOT EXISTS idx_export_jobs_format ON arkham_export_jobs(format)
        """)
        await self._db.execute("""
            CREATE INDEX IF NOT EXISTS idx_export_jobs_target ON arkham_export_jobs(target)
        """)

        # ===========================================
        # Multi-tenancy Migration
        # ===========================================
        await self._db.execute("""
            DO $$
            DECLARE
                tables_to_update TEXT[] := ARRAY['arkham_export_jobs'];
                tbl TEXT;
            BEGIN
                FOREACH tbl IN ARRAY tables_to_update LOOP
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_schema = 'public'
                        AND table_name = tbl
                        AND column_name = 'tenant_id'
                    ) THEN
                        EXECUTE format('ALTER TABLE %I ADD COLUMN tenant_id UUID', tbl);
                    END IF;
                END LOOP;
            END $$;
        """)

        await self._db.execute("""
            CREATE INDEX IF NOT EXISTS idx_export_jobs_tenant
            ON arkham_export_jobs(tenant_id)
        """)

        logger.debug("Export schema created/verified")

    # === Public API Methods ===

    async def create_export_job(
        self,
        format: ExportFormat,
        target: ExportTarget,
        filters: Optional[Dict[str, Any]] = None,
        options: Optional[ExportOptions] = None,
        created_by: str = "system",
    ) -> ExportJob:
        """Create a new export job."""
        job_id = str(uuid4())
        now = datetime.utcnow()

        # Default options
        if options is None:
            options = ExportOptions()

        # Calculate expiration (24 hours from creation)
        expires_at = now + timedelta(hours=24)

        job = ExportJob(
            id=job_id,
            format=format,
            target=target,
            status=ExportStatus.PENDING,
            created_at=now,
            filters=filters or {},
            options=options,
            created_by=created_by,
            expires_at=expires_at,
        )

        await self._save_job(job)

        # Emit event
        if self._events:
            await self._events.emit(
                "export.job.created",
                {
                    "job_id": job_id,
                    "format": format.value,
                    "target": target.value,
                    "created_by": created_by,
                },
                source=self.name,
            )

        # Start processing asynchronously (in real implementation)
        # For now, we'll process immediately
        await self._process_job(job)

        return job

    async def get_job_status(self, job_id: str) -> Optional[ExportJob]:
        """Get the status of an export job."""
        if not self._db:
            return None

        # Filter by tenant_id for multi-tenancy
        tenant_id = self.get_tenant_id_or_none()
        if tenant_id:
            row = await self._db.fetch_one(
                "SELECT * FROM arkham_export_jobs WHERE id = ? AND tenant_id = ?",
                [job_id, str(tenant_id)],
            )
        else:
            row = await self._db.fetch_one(
                "SELECT * FROM arkham_export_jobs WHERE id = ?",
                [job_id],
            )
        return self._row_to_job(row) if row else None

    async def list_jobs(
        self,
        filter: Optional[ExportFilter] = None,
        limit: int = 50,
        offset: int = 0,
    ) -> List[ExportJob]:
        """List export jobs with optional filtering."""
        if not self._db:
            return []

        query = "SELECT * FROM arkham_export_jobs WHERE 1=1"
        params = []

        # Filter by tenant_id for multi-tenancy
        tenant_id = self.get_tenant_id_or_none()
        if tenant_id:
            query += " AND tenant_id = ?"
            params.append(str(tenant_id))

        if filter:
            if filter.status:
                query += " AND status = ?"
                params.append(filter.status.value)
            if filter.format:
                query += " AND format = ?"
                params.append(filter.format.value)
            if filter.target:
                query += " AND target = ?"
                params.append(filter.target.value)
            if filter.created_by:
                query += " AND created_by = ?"
                params.append(filter.created_by)
            if filter.created_after:
                query += " AND created_at >= ?"
                params.append(filter.created_after.isoformat())
            if filter.created_before:
                query += " AND created_at <= ?"
                params.append(filter.created_before.isoformat())

        query += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
        params.extend([limit, offset])

        rows = await self._db.fetch_all(query, params)
        return [self._row_to_job(row) for row in rows]

    async def cancel_job(self, job_id: str) -> Optional[ExportJob]:
        """Cancel a pending export job."""
        job = await self.get_job_status(job_id)
        if not job:
            return None

        if job.status in [ExportStatus.PENDING, ExportStatus.PROCESSING]:
            job.status = ExportStatus.CANCELLED
            job.completed_at = datetime.utcnow()
            await self._save_job(job, update=True)

            # Emit event
            if self._events:
                await self._events.emit(
                    "export.job.cancelled",
                    {"job_id": job_id},
                    source=self.name,
                )

        return job

    async def get_download_url(self, job_id: str) -> Optional[str]:
        """Get download URL for a completed export job."""
        job = await self.get_job_status(job_id)
        if not job or job.status != ExportStatus.COMPLETED:
            return None

        # Check if file expired
        if job.expires_at and datetime.utcnow() > job.expires_at:
            return None

        return job.download_url

    async def get_statistics(self) -> ExportStatistics:
        """Get statistics about export jobs."""
        if not self._db:
            return ExportStatistics()

        # Filter by tenant_id for multi-tenancy
        tenant_id = self.get_tenant_id_or_none()
        tenant_filter = ""
        tenant_params = []
        if tenant_id:
            tenant_filter = " WHERE tenant_id = ?"
            tenant_params = [str(tenant_id)]

        # Total jobs
        total = await self._db.fetch_one(
            f"SELECT COUNT(*) as count FROM arkham_export_jobs{tenant_filter}",
            tenant_params,
        )
        total_jobs = total["count"] if total else 0

        # By status
        status_rows = await self._db.fetch_all(
            f"SELECT status, COUNT(*) as count FROM arkham_export_jobs{tenant_filter} GROUP BY status",
            tenant_params,
        )
        by_status = {row["status"]: row["count"] for row in status_rows}

        # By format
        format_rows = await self._db.fetch_all(
            f"SELECT format, COUNT(*) as count FROM arkham_export_jobs{tenant_filter} GROUP BY format",
            tenant_params,
        )
        by_format = {row["format"]: row["count"] for row in format_rows}

        # By target
        target_rows = await self._db.fetch_all(
            f"SELECT target, COUNT(*) as count FROM arkham_export_jobs{tenant_filter} GROUP BY target",
            tenant_params,
        )
        by_target = {row["target"]: row["count"] for row in target_rows}

        # Aggregates
        if tenant_id:
            totals = await self._db.fetch_one(
                """
                SELECT
                    SUM(record_count) as total_records,
                    SUM(file_size) as total_size,
                    AVG(processing_time_ms) as avg_time
                FROM arkham_export_jobs
                WHERE status = 'completed' AND tenant_id = ?
                """,
                [str(tenant_id)],
            )
        else:
            totals = await self._db.fetch_one("""
                SELECT
                    SUM(record_count) as total_records,
                    SUM(file_size) as total_size,
                    AVG(processing_time_ms) as avg_time
                FROM arkham_export_jobs
                WHERE status = 'completed'
            """)

        # Oldest pending
        if tenant_id:
            oldest = await self._db.fetch_one(
                """
                SELECT MIN(created_at) as oldest
                FROM arkham_export_jobs
                WHERE status = 'pending' AND tenant_id = ?
                """,
                [str(tenant_id)],
            )
        else:
            oldest = await self._db.fetch_one("""
                SELECT MIN(created_at) as oldest
                FROM arkham_export_jobs
                WHERE status = 'pending'
            """)

        return ExportStatistics(
            total_jobs=total_jobs,
            by_status=by_status,
            by_format=by_format,
            by_target=by_target,
            jobs_pending=by_status.get("pending", 0),
            jobs_processing=by_status.get("processing", 0),
            jobs_completed=by_status.get("completed", 0),
            jobs_failed=by_status.get("failed", 0),
            total_records_exported=totals["total_records"] if totals and totals["total_records"] else 0,
            total_file_size_bytes=totals["total_size"] if totals and totals["total_size"] else 0,
            avg_processing_time_ms=totals["avg_time"] if totals and totals["avg_time"] else 0.0,
            oldest_pending_job=datetime.fromisoformat(oldest["oldest"]) if oldest and oldest["oldest"] else None,
        )

    async def get_count(self, status: Optional[str] = None) -> int:
        """Get count of export jobs, optionally filtered by status."""
        if not self._db:
            return 0

        # Filter by tenant_id for multi-tenancy
        tenant_id = self.get_tenant_id_or_none()

        if status:
            if tenant_id:
                result = await self._db.fetch_one(
                    "SELECT COUNT(*) as count FROM arkham_export_jobs WHERE status = ? AND tenant_id = ?",
                    [status, str(tenant_id)],
                )
            else:
                result = await self._db.fetch_one(
                    "SELECT COUNT(*) as count FROM arkham_export_jobs WHERE status = ?",
                    [status],
                )
        else:
            if tenant_id:
                result = await self._db.fetch_one(
                    "SELECT COUNT(*) as count FROM arkham_export_jobs WHERE status = 'pending' AND tenant_id = ?",
                    [str(tenant_id)],
                )
            else:
                result = await self._db.fetch_one(
                    "SELECT COUNT(*) as count FROM arkham_export_jobs WHERE status = 'pending'"
                )

        return result["count"] if result else 0

    def get_supported_formats(self) -> List[FormatInfo]:
        """Get list of supported export formats."""
        return [
            FormatInfo(
                format=ExportFormat.JSON,
                name="JSON",
                description="JavaScript Object Notation - structured data",
                file_extension=".json",
                mime_type="application/json",
                supports_metadata=True,
            ),
            FormatInfo(
                format=ExportFormat.CSV,
                name="CSV",
                description="Comma-Separated Values - tabular data",
                file_extension=".csv",
                mime_type="text/csv",
                supports_flatten=True,
            ),
            FormatInfo(
                format=ExportFormat.PDF,
                name="PDF",
                description="Portable Document Format - formatted documents",
                file_extension=".pdf",
                mime_type="application/pdf",
                placeholder=False,  # Now implemented
            ),
            FormatInfo(
                format=ExportFormat.DOCX,
                name="DOCX",
                description="Microsoft Word Document",
                file_extension=".docx",
                mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                placeholder=True,  # Not yet implemented
            ),
            FormatInfo(
                format=ExportFormat.XLSX,
                name="XLSX",
                description="Microsoft Excel Spreadsheet",
                file_extension=".xlsx",
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                placeholder=False,  # Now implemented
            ),
        ]

    def get_export_targets(self) -> List[TargetInfo]:
        """Get list of available export targets."""
        return [
            TargetInfo(
                target=ExportTarget.DOCUMENTS,
                name="Documents",
                description="Export document records with text and metadata",
                available_formats=[ExportFormat.JSON, ExportFormat.CSV, ExportFormat.XLSX],
            ),
            TargetInfo(
                target=ExportTarget.ENTITIES,
                name="Entities",
                description="Export extracted entities with relationships",
                available_formats=[ExportFormat.JSON, ExportFormat.CSV, ExportFormat.XLSX],
            ),
            TargetInfo(
                target=ExportTarget.CLAIMS,
                name="Claims",
                description="Export claims with evidence and verification status",
                available_formats=[ExportFormat.JSON, ExportFormat.CSV, ExportFormat.PDF, ExportFormat.XLSX],
            ),
            TargetInfo(
                target=ExportTarget.TIMELINE,
                name="Timeline",
                description="Export timeline events in chronological order",
                available_formats=[ExportFormat.JSON, ExportFormat.CSV, ExportFormat.XLSX],
            ),
            TargetInfo(
                target=ExportTarget.GRAPH,
                name="Graph",
                description="Export graph nodes and edges",
                available_formats=[ExportFormat.JSON],
            ),
            TargetInfo(
                target=ExportTarget.MATRIX,
                name="ACH Matrix",
                description="Export ACH matrix with hypotheses and evidence",
                available_formats=[ExportFormat.JSON, ExportFormat.CSV, ExportFormat.PDF, ExportFormat.XLSX],
            ),
        ]

    # === Private Helper Methods ===

    async def _process_job(self, job: ExportJob) -> ExportResult:
        """Process an export job."""
        import time
        start_time = time.time()

        try:
            job.status = ExportStatus.PROCESSING
            job.started_at = datetime.utcnow()
            await self._save_job(job, update=True)

            # Emit started event
            if self._events:
                await self._events.emit(
                    "export.job.started",
                    {"job_id": job.id, "format": job.format.value, "target": job.target.value},
                    source=self.name,
                )

            # Generate export file with real data
            file_path, record_count = await self._generate_export_file(job)

            # Update job
            job.status = ExportStatus.COMPLETED
            job.completed_at = datetime.utcnow()
            job.file_path = file_path
            job.file_size = os.path.getsize(file_path) if file_path and os.path.exists(file_path) else 0
            job.download_url = f"/api/export/jobs/{job.id}/download"
            job.record_count = record_count
            job.processing_time_ms = (time.time() - start_time) * 1000

            await self._save_job(job, update=True)

            # Emit completed event
            if self._events:
                await self._events.emit(
                    "export.job.completed",
                    {
                        "job_id": job.id,
                        "record_count": record_count,
                        "file_size": job.file_size,
                        "processing_time_ms": job.processing_time_ms,
                    },
                    source=self.name,
                )
                await self._events.emit(
                    "export.file.created",
                    {
                        "job_id": job.id,
                        "file_path": file_path,
                        "file_size": job.file_size,
                    },
                    source=self.name,
                )

            return ExportResult(
                job_id=job.id,
                success=True,
                file_path=file_path,
                file_size=job.file_size,
                download_url=job.download_url,
                expires_at=job.expires_at,
                record_count=record_count,
                processing_time_ms=job.processing_time_ms,
            )

        except Exception as e:
            logger.error(f"Export job {job.id} failed: {e}")
            job.status = ExportStatus.FAILED
            job.error = str(e)
            job.completed_at = datetime.utcnow()
            await self._save_job(job, update=True)

            # Emit failed event
            if self._events:
                await self._events.emit(
                    "export.job.failed",
                    {"job_id": job.id, "error": str(e)},
                    source=self.name,
                )

            return ExportResult(
                job_id=job.id,
                success=False,
                error=str(e),
                processing_time_ms=(time.time() - start_time) * 1000,
            )

    async def _generate_export_file(self, job: ExportJob) -> tuple[Optional[str], int]:
        """Generate export file with real data."""
        # Fetch data based on target
        data = await self._fetch_data(job.target, job.filters, job.options)
        # Calculate record count - handle timeline dict format
        if isinstance(data, list):
            record_count = len(data)
        elif isinstance(data, dict) and "events" in data:
            # Timeline format with events list
            record_count = len(data.get("events", []))
        else:
            record_count = 1

        # Generate file based on format
        filename = f"{job.id}_{job.target.value}.{self._get_file_extension(job.format)}"
        file_path = os.path.join(self._export_dir, filename)

        if job.format == ExportFormat.JSON:
            await self._generate_json(file_path, data, job)
        elif job.format == ExportFormat.CSV:
            await self._generate_csv(file_path, data, job)
        elif job.format == ExportFormat.PDF:
            await self._generate_pdf(file_path, data, job)
        elif job.format == ExportFormat.XLSX:
            await self._generate_xlsx(file_path, data, job)
        elif job.format == ExportFormat.DOCX:
            # DOCX not yet implemented
            with open(file_path, "wb") as f:
                f.write(b"DOCX export not yet implemented")
            return file_path, 0

        return file_path, record_count

    async def _fetch_data(
        self,
        target: ExportTarget,
        filters: Dict[str, Any],
        options: Optional[ExportOptions],
    ) -> List[Dict[str, Any]]:
        """Fetch data from appropriate shard API."""
        try:
            async with httpx.AsyncClient(timeout=60.0) as client:
                if target == ExportTarget.DOCUMENTS:
                    return await self._fetch_documents(client, filters, options)
                elif target == ExportTarget.ENTITIES:
                    return await self._fetch_entities(client, filters, options)
                elif target == ExportTarget.CLAIMS:
                    return await self._fetch_claims(client, filters, options)
                elif target == ExportTarget.TIMELINE:
                    return await self._fetch_timeline(client, filters, options)
                elif target == ExportTarget.GRAPH:
                    return await self._fetch_graph(client, filters, options)
                elif target == ExportTarget.MATRIX:
                    return await self._fetch_ach_matrices(client, filters, options)
                else:
                    logger.warning(f"Unknown export target: {target}")
                    return []
        except Exception as e:
            logger.error(f"Failed to fetch data for {target}: {e}")
            raise

    async def _fetch_documents(
        self,
        client: httpx.AsyncClient,
        filters: Dict[str, Any],
        options: Optional[ExportOptions],
    ) -> List[Dict[str, Any]]:
        """Fetch documents from the documents shard."""
        documents = []
        page = 1
        page_size = 100
        max_records = options.max_records if options else None

        while True:
            params = {"page": page, "page_size": page_size}
            if filters.get("project_id"):
                params["project_id"] = filters["project_id"]
            if filters.get("status"):
                params["status"] = filters["status"]

            response = await client.get(f"{INTERNAL_API_BASE}/api/documents/items", params=params)
            if response.status_code != 200:
                logger.error(f"Documents API error: {response.status_code}")
                break

            data = response.json()
            items = data.get("items", [])
            if not items:
                break

            documents.extend(items)

            # Check if we've reached max records
            if max_records and len(documents) >= max_records:
                documents = documents[:max_records]
                break

            # Check if there are more pages
            if len(items) < page_size:
                break

            page += 1

        return documents

    async def _fetch_entities(
        self,
        client: httpx.AsyncClient,
        filters: Dict[str, Any],
        options: Optional[ExportOptions],
    ) -> List[Dict[str, Any]]:
        """Fetch entities from the entities shard."""
        entities = []
        page = 1
        page_size = 100
        max_records = options.max_records if options else None

        while True:
            params = {"page": page, "page_size": page_size}
            if filters.get("entity_type"):
                params["filter"] = filters["entity_type"]
            if options and options.entity_types:
                params["filter"] = options.entity_types[0]  # API only supports one type

            response = await client.get(f"{INTERNAL_API_BASE}/api/entities/items", params=params)
            if response.status_code != 200:
                logger.error(f"Entities API error: {response.status_code}")
                break

            data = response.json()
            items = data.get("items", [])
            if not items:
                break

            entities.extend(items)

            if max_records and len(entities) >= max_records:
                entities = entities[:max_records]
                break

            if len(items) < page_size:
                break

            page += 1

        # Optionally fetch relationships for each entity
        if options and options.include_relationships:
            for entity in entities:
                try:
                    rel_response = await client.get(
                        f"{INTERNAL_API_BASE}/api/entities/{entity['id']}/relationships"
                    )
                    if rel_response.status_code == 200:
                        entity["relationships"] = rel_response.json()
                except Exception as e:
                    logger.warning(f"Failed to fetch relationships for entity {entity['id']}: {e}")

        return entities

    async def _fetch_claims(
        self,
        client: httpx.AsyncClient,
        filters: Dict[str, Any],
        options: Optional[ExportOptions],
    ) -> List[Dict[str, Any]]:
        """Fetch claims from the claims shard."""
        claims = []
        page = 1
        page_size = 100
        max_records = options.max_records if options else None

        while True:
            params = {"page": page, "page_size": page_size}
            if filters.get("status"):
                params["status"] = filters["status"]
            if filters.get("document_id"):
                params["document_id"] = filters["document_id"]

            response = await client.get(f"{INTERNAL_API_BASE}/api/claims/", params=params)
            if response.status_code != 200:
                logger.error(f"Claims API error: {response.status_code}")
                break

            data = response.json()
            items = data.get("items", [])
            if not items:
                break

            claims.extend(items)

            if max_records and len(claims) >= max_records:
                claims = claims[:max_records]
                break

            if len(items) < page_size:
                break

            page += 1

        # Fetch evidence for each claim if needed
        if options and options.include_relationships:
            for claim in claims:
                try:
                    ev_response = await client.get(
                        f"{INTERNAL_API_BASE}/api/claims/{claim['id']}/evidence"
                    )
                    if ev_response.status_code == 200:
                        claim["evidence"] = ev_response.json()
                except Exception as e:
                    logger.warning(f"Failed to fetch evidence for claim {claim['id']}: {e}")

        return claims

    async def _fetch_timeline(
        self,
        client: httpx.AsyncClient,
        filters: Dict[str, Any],
        options: Optional[ExportOptions],
    ) -> Dict[str, Any]:
        """Fetch timeline events with optional gaps, conflicts, and entity info."""
        params = {"limit": 1000, "offset": 0}

        if options:
            if options.date_range_start:
                params["start_date"] = options.date_range_start.isoformat()
            if options.date_range_end:
                params["end_date"] = options.date_range_end.isoformat()

        max_records = options.max_records if options else None
        if max_records:
            params["limit"] = min(max_records, 1000)

        # Fetch events
        response = await client.get(f"{INTERNAL_API_BASE}/api/timeline/events", params=params)
        if response.status_code != 200:
            logger.error(f"Timeline API error: {response.status_code}")
            return {"events": [], "gaps": [], "conflicts": [], "entities": {}}

        data = response.json()
        events = data.get("events", [])

        result = {
            "events": events,
            "gaps": [],
            "conflicts": [],
            "entities": {},
            "stats": {
                "total_events": len(events),
                "date_range": {
                    "start": events[0].get("date_start") if events else None,
                    "end": events[-1].get("date_start") if events else None,
                },
            },
        }

        # Fetch gaps if requested
        if options and options.include_gaps:
            try:
                gap_params = {}
                if options.date_range_start:
                    gap_params["start_date"] = options.date_range_start.isoformat()
                if options.date_range_end:
                    gap_params["end_date"] = options.date_range_end.isoformat()
                gaps_response = await client.get(f"{INTERNAL_API_BASE}/api/timeline/gaps", params=gap_params)
                if gaps_response.status_code == 200:
                    gaps_data = gaps_response.json()
                    result["gaps"] = gaps_data.get("gaps", [])
                    result["stats"]["gap_count"] = len(result["gaps"])
            except Exception as e:
                logger.warning(f"Failed to fetch timeline gaps: {e}")

        # Fetch conflicts if requested
        if options and options.include_conflicts:
            try:
                conflicts_response = await client.get(f"{INTERNAL_API_BASE}/api/timeline/conflicts/analyze")
                if conflicts_response.status_code == 200:
                    conflicts_data = conflicts_response.json()
                    result["conflicts"] = conflicts_data.get("conflicts", [])
                    result["stats"]["conflict_count"] = len(result["conflicts"])
            except Exception as e:
                logger.warning(f"Failed to fetch timeline conflicts: {e}")

        # Fetch entity names if events have entity IDs
        entity_ids = set()
        for event in events:
            for eid in event.get("entities", []):
                entity_ids.add(eid)

        if entity_ids:
            try:
                for eid in list(entity_ids)[:50]:  # Limit to 50 entities
                    ent_response = await client.get(f"{INTERNAL_API_BASE}/api/entities/{eid}")
                    if ent_response.status_code == 200:
                        ent_data = ent_response.json()
                        result["entities"][eid] = {
                            "name": ent_data.get("name", "Unknown"),
                            "type": ent_data.get("entity_type", "UNKNOWN"),
                        }
            except Exception as e:
                logger.warning(f"Failed to fetch entity names: {e}")

        # Group events if requested
        if options and options.group_by:
            result["grouped_events"] = self._group_timeline_events(events, options.group_by, result.get("entities", {}))

        return result

    def _group_timeline_events(
        self,
        events: List[Dict[str, Any]],
        group_by: str,
        entities: Dict[str, Dict[str, str]],
    ) -> Dict[str, List[Dict[str, Any]]]:
        """Group timeline events by day/week/month/entity."""
        from collections import defaultdict
        grouped = defaultdict(list)

        for event in events:
            date_str = event.get("date_start", "")
            if not date_str:
                grouped["Unknown"].append(event)
                continue

            try:
                # Parse date - handle various formats
                if "T" in date_str:
                    dt = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
                else:
                    dt = datetime.strptime(date_str[:10], "%Y-%m-%d")

                if group_by == "day":
                    key = dt.strftime("%Y-%m-%d")
                elif group_by == "week":
                    key = f"{dt.year}-W{dt.isocalendar()[1]:02d}"
                elif group_by == "month":
                    key = dt.strftime("%Y-%m")
                elif group_by == "entity":
                    # Group by entities mentioned in the event
                    event_entities = event.get("entities", [])
                    if event_entities:
                        for eid in event_entities:
                            ent_name = entities.get(eid, {}).get("name", eid[:8])
                            grouped[ent_name].append(event)
                    else:
                        grouped["No Entity"].append(event)
                    continue
                else:
                    key = "Unknown"

                grouped[key].append(event)
            except Exception:
                grouped["Unknown"].append(event)

        return dict(grouped)

    async def _fetch_graph(
        self,
        client: httpx.AsyncClient,
        filters: Dict[str, Any],
        options: Optional[ExportOptions],
    ) -> Dict[str, Any]:
        """Fetch graph data from the graph shard."""
        project_id = filters.get("project_id", "default")

        response = await client.get(f"{INTERNAL_API_BASE}/api/graph/{project_id}")
        if response.status_code != 200:
            logger.error(f"Graph API error: {response.status_code}")
            return {"nodes": [], "edges": []}

        return response.json()

    async def _fetch_ach_matrices(
        self,
        client: httpx.AsyncClient,
        filters: Dict[str, Any],
        options: Optional[ExportOptions],
    ) -> List[Dict[str, Any]]:
        """Fetch ACH matrices from the ACH shard."""
        params = {}
        if filters.get("project_id"):
            params["project_id"] = filters["project_id"]

        response = await client.get(f"{INTERNAL_API_BASE}/api/ach/matrices", params=params)
        if response.status_code != 200:
            logger.error(f"ACH API error: {response.status_code}")
            return []

        data = response.json()
        matrices = data.get("matrices", [])

        # Fetch full details for each matrix
        detailed_matrices = []
        for matrix_summary in matrices:
            try:
                detail_response = await client.get(
                    f"{INTERNAL_API_BASE}/api/ach/matrix/{matrix_summary['id']}"
                )
                if detail_response.status_code == 200:
                    detailed_matrices.append(detail_response.json())
                else:
                    detailed_matrices.append(matrix_summary)
            except Exception as e:
                logger.warning(f"Failed to fetch ACH matrix {matrix_summary['id']}: {e}")
                detailed_matrices.append(matrix_summary)

        return detailed_matrices

    # === Format Generators ===

    async def _generate_json(
        self,
        file_path: str,
        data: Any,
        job: ExportJob,
    ) -> None:
        """Generate JSON export file."""
        output = {
            "export_info": {
                "job_id": job.id,
                "target": job.target.value,
                "exported_at": datetime.utcnow().isoformat(),
                "record_count": len(data) if isinstance(data, list) else 1,
            },
            "data": data,
        }

        if job.options and job.options.include_metadata:
            output["metadata"] = {
                "filters": job.filters,
                "options": {
                    "include_metadata": job.options.include_metadata,
                    "include_relationships": job.options.include_relationships,
                },
            }

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(output, f, indent=2, default=str, ensure_ascii=False)

    async def _generate_csv(
        self,
        file_path: str,
        data: Any,
        job: ExportJob,
    ) -> None:
        """Generate CSV export file."""
        if not isinstance(data, list) or not data:
            # Write empty CSV with header comment
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("# No data to export\n")
            return

        # Flatten nested structures if needed
        flat_data = self._flatten_for_csv(data)

        # Get all unique keys across all records
        all_keys = set()
        for record in flat_data:
            all_keys.update(record.keys())
        fieldnames = sorted(all_keys)

        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(flat_data)

    def _flatten_for_csv(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Flatten nested structures for CSV export."""
        flattened = []
        for record in data:
            flat_record = {}
            for key, value in record.items():
                if isinstance(value, (list, dict)):
                    # Convert complex types to JSON strings
                    flat_record[key] = json.dumps(value, default=str)
                else:
                    flat_record[key] = value
            flattened.append(flat_record)
        return flattened

    async def _generate_pdf(
        self,
        file_path: str,
        data: Any,
        job: ExportJob,
    ) -> None:
        """Generate PDF export file using reportlab."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                SimpleDocTemplate,
                Paragraph,
                Spacer,
                Table,
                TableStyle,
                PageBreak,
            )
        except ImportError:
            logger.error("reportlab not installed")
            with open(file_path, "wb") as f:
                f.write(b"PDF generation requires reportlab library")
            return

        doc = SimpleDocTemplate(file_path, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=20,
        )
        story.append(Paragraph(f"Export: {job.target.value.title()}", title_style))
        story.append(Spacer(1, 12))

        # Export info
        info_style = styles["Normal"]
        story.append(Paragraph(f"<b>Export Date:</b> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}", info_style))
        story.append(Paragraph(f"<b>Job ID:</b> {job.id[:8]}...", info_style))
        record_count = len(data) if isinstance(data, list) else 1
        story.append(Paragraph(f"<b>Records:</b> {record_count}", info_style))
        story.append(Spacer(1, 20))

        # Content based on target type
        if job.target == ExportTarget.CLAIMS:
            self._add_claims_to_pdf(story, data, styles)
        elif job.target == ExportTarget.MATRIX:
            self._add_matrices_to_pdf(story, data, styles)
        elif job.target == ExportTarget.DOCUMENTS:
            self._add_documents_to_pdf(story, data, styles)
        elif job.target == ExportTarget.ENTITIES:
            self._add_entities_to_pdf(story, data, styles)
        elif job.target == ExportTarget.TIMELINE:
            self._add_timeline_to_pdf(story, data, styles)
        else:
            # Generic list export
            self._add_generic_list_to_pdf(story, data, styles)

        doc.build(story)

    def _add_claims_to_pdf(self, story, data, styles):
        """Add claims to PDF."""
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors

        heading_style = styles["Heading2"]
        normal_style = styles["Normal"]

        if not isinstance(data, list):
            data = [data]

        for i, claim in enumerate(data):
            story.append(Paragraph(f"Claim {i + 1}", heading_style))
            story.append(Spacer(1, 6))

            # Claim text
            claim_text = claim.get("text", "N/A")
            story.append(Paragraph(f"<b>Text:</b> {claim_text[:500]}...", normal_style))

            # Status and type
            story.append(Paragraph(f"<b>Status:</b> {claim.get('status', 'N/A')}", normal_style))
            story.append(Paragraph(f"<b>Type:</b> {claim.get('claim_type', 'N/A')}", normal_style))
            story.append(Paragraph(f"<b>Confidence:</b> {claim.get('confidence', 0):.2f}", normal_style))

            # Evidence counts
            story.append(Paragraph(
                f"<b>Evidence:</b> {claim.get('evidence_count', 0)} total "
                f"({claim.get('supporting_count', 0)} supporting, {claim.get('refuting_count', 0)} refuting)",
                normal_style
            ))

            story.append(Spacer(1, 12))

    def _add_matrices_to_pdf(self, story, data, styles):
        """Add ACH matrices to PDF."""
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors

        heading_style = styles["Heading2"]
        normal_style = styles["Normal"]

        if not isinstance(data, list):
            data = [data]

        for matrix in data:
            story.append(Paragraph(f"Matrix: {matrix.get('title', 'Untitled')}", heading_style))
            story.append(Spacer(1, 6))

            story.append(Paragraph(f"<b>Description:</b> {matrix.get('description', 'N/A')}", normal_style))
            story.append(Paragraph(f"<b>Status:</b> {matrix.get('status', 'N/A')}", normal_style))

            # Hypotheses
            hypotheses = matrix.get("hypotheses", [])
            if hypotheses:
                story.append(Spacer(1, 6))
                story.append(Paragraph("<b>Hypotheses:</b>", normal_style))
                for h in hypotheses:
                    story.append(Paragraph(f"  - {h.get('title', 'Untitled')}", normal_style))

            # Evidence
            evidence = matrix.get("evidence", [])
            if evidence:
                story.append(Spacer(1, 6))
                story.append(Paragraph(f"<b>Evidence Items:</b> {len(evidence)}", normal_style))

            story.append(Spacer(1, 12))

    def _add_documents_to_pdf(self, story, data, styles):
        """Add documents to PDF."""
        from reportlab.platypus import Paragraph, Spacer

        heading_style = styles["Heading2"]
        normal_style = styles["Normal"]

        if not isinstance(data, list):
            data = [data]

        for i, doc in enumerate(data[:50]):  # Limit to 50 for PDF
            story.append(Paragraph(f"Document {i + 1}: {doc.get('title', doc.get('filename', 'Untitled'))}", heading_style))
            story.append(Spacer(1, 6))

            story.append(Paragraph(f"<b>Filename:</b> {doc.get('filename', 'N/A')}", normal_style))
            story.append(Paragraph(f"<b>Type:</b> {doc.get('file_type', 'N/A')}", normal_style))
            story.append(Paragraph(f"<b>Status:</b> {doc.get('status', 'N/A')}", normal_style))
            story.append(Paragraph(f"<b>Pages:</b> {doc.get('page_count', 0)}", normal_style))
            story.append(Paragraph(f"<b>Entities:</b> {doc.get('entity_count', 0)}", normal_style))

            story.append(Spacer(1, 12))

    def _add_entities_to_pdf(self, story, data, styles):
        """Add entities to PDF."""
        from reportlab.platypus import Paragraph, Spacer

        heading_style = styles["Heading2"]
        normal_style = styles["Normal"]

        if not isinstance(data, list):
            data = [data]

        # Group by entity type
        by_type = {}
        for entity in data:
            etype = entity.get("entity_type", "UNKNOWN")
            if etype not in by_type:
                by_type[etype] = []
            by_type[etype].append(entity)

        for etype, entities in by_type.items():
            story.append(Paragraph(f"{etype} ({len(entities)} entities)", heading_style))
            story.append(Spacer(1, 6))

            for entity in entities[:20]:  # Limit per type
                story.append(Paragraph(f"  - {entity.get('name', 'N/A')}", normal_style))

            if len(entities) > 20:
                story.append(Paragraph(f"  ... and {len(entities) - 20} more", normal_style))

            story.append(Spacer(1, 12))

    def _add_timeline_to_pdf(self, story, data, styles):
        """Add timeline events to PDF with enhanced formatting."""
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle

        heading_style = styles["Heading2"]
        heading3_style = ParagraphStyle(
            "Heading3",
            parent=styles["Heading3"] if "Heading3" in styles else styles["Heading2"],
            fontSize=11,
            spaceAfter=6,
        )
        normal_style = styles["Normal"]

        # Handle new format (dict with events, gaps, conflicts, entities)
        if isinstance(data, dict):
            events = data.get("events", [])
            gaps = data.get("gaps", [])
            conflicts = data.get("conflicts", [])
            entities = data.get("entities", {})
            stats = data.get("stats", {})
            grouped_events = data.get("grouped_events")
        else:
            # Legacy format - list of events
            events = data if isinstance(data, list) else [data]
            gaps = []
            conflicts = []
            entities = {}
            stats = {}
            grouped_events = None

        # Summary statistics
        story.append(Paragraph("Timeline Summary", heading_style))
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"<b>Total Events:</b> {len(events)}", normal_style))
        if stats.get("date_range"):
            dr = stats["date_range"]
            story.append(Paragraph(f"<b>Date Range:</b> {dr.get('start', 'N/A')} to {dr.get('end', 'N/A')}", normal_style))
        if stats.get("gap_count"):
            story.append(Paragraph(f"<b>Timeline Gaps:</b> {stats['gap_count']}", normal_style))
        if stats.get("conflict_count"):
            story.append(Paragraph(f"<b>Conflicts Detected:</b> {stats['conflict_count']}", normal_style))
        story.append(Spacer(1, 12))

        # Conflicts section (if any)
        if conflicts:
            story.append(Paragraph("Conflicts / Issues", heading_style))
            story.append(Spacer(1, 6))
            for i, conflict in enumerate(conflicts[:20]):
                conflict_type = conflict.get("conflict_type", "Unknown")
                severity = conflict.get("severity", "medium")
                desc = conflict.get("description", "")[:150]
                story.append(Paragraph(
                    f"<b>{i+1}. [{severity.upper()}] {conflict_type}:</b> {desc}",
                    normal_style
                ))
                if conflict.get("resolution_suggestion"):
                    story.append(Paragraph(
                        f"   <i>Suggestion: {conflict['resolution_suggestion'][:100]}</i>",
                        normal_style
                    ))
            story.append(Spacer(1, 12))

        # Gaps section (if any)
        if gaps:
            story.append(Paragraph("Timeline Gaps", heading_style))
            story.append(Spacer(1, 6))
            for gap in gaps[:15]:
                start = gap.get("gap_start", "?")
                end = gap.get("gap_end", "?")
                days = gap.get("days", 0)
                severity = gap.get("severity", "low")
                story.append(Paragraph(
                    f"<b>[{severity.upper()}]</b> {start} to {end} ({days} days)",
                    normal_style
                ))
            story.append(Spacer(1, 12))

        # Events section
        story.append(Paragraph("Timeline Events", heading_style))
        story.append(Spacer(1, 6))

        # If grouped, show by group
        if grouped_events:
            for group_name, group_events in sorted(grouped_events.items()):
                story.append(Paragraph(f"<b>{group_name}</b> ({len(group_events)} events)", heading3_style))
                for event in group_events[:20]:
                    self._add_single_event_to_pdf(story, event, normal_style, entities)
                if len(group_events) > 20:
                    story.append(Paragraph(f"  ... and {len(group_events) - 20} more events", normal_style))
                story.append(Spacer(1, 8))
        else:
            # Chronological list
            for event in events[:100]:
                self._add_single_event_to_pdf(story, event, normal_style, entities)

        if len(events) > 100 and not grouped_events:
            story.append(Paragraph(f"<i>... and {len(events) - 100} more events (truncated for PDF)</i>", normal_style))

    def _add_single_event_to_pdf(self, story, event, style, entities):
        """Add a single timeline event to the PDF."""
        from reportlab.platypus import Paragraph, Spacer

        date_str = event.get("date_start", "Unknown date")
        precision = event.get("precision", "")
        text = event.get("text", "No description")[:250]
        event_type = event.get("event_type", "")
        source = event.get("source_document", "")

        # Format date with precision indicator
        if precision and precision != "day":
            date_str = f"{date_str} ({precision})"

        # Build event line
        line = f"<b>{date_str}</b>"
        if event_type:
            line += f" [{event_type}]"
        line += f": {text}"

        story.append(Paragraph(line, style))

        # Add entity names if present
        event_entities = event.get("entities", [])
        if event_entities and entities:
            entity_names = []
            for eid in event_entities[:5]:
                ent_info = entities.get(eid, {})
                name = ent_info.get("name", eid[:8])
                entity_names.append(name)
            if entity_names:
                story.append(Paragraph(f"   <i>Entities: {', '.join(entity_names)}</i>", style))

        # Add source if present
        if source:
            story.append(Paragraph(f"   <i>Source: {source[:50]}</i>", style))

        story.append(Spacer(1, 4))

    def _add_generic_list_to_pdf(self, story, data, styles):
        """Add generic list data to PDF."""
        from reportlab.platypus import Paragraph, Spacer

        normal_style = styles["Normal"]

        if not isinstance(data, list):
            data = [data]

        for i, item in enumerate(data[:100]):  # Limit
            if isinstance(item, dict):
                story.append(Paragraph(f"<b>Item {i + 1}:</b>", normal_style))
                for key, value in item.items():
                    if not isinstance(value, (list, dict)):
                        story.append(Paragraph(f"  {key}: {value}", normal_style))
                story.append(Spacer(1, 6))

    async def _generate_xlsx(
        self,
        file_path: str,
        data: Any,
        job: ExportJob,
    ) -> None:
        """Generate XLSX export file using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            logger.error("openpyxl not installed")
            with open(file_path, "wb") as f:
                f.write(b"XLSX generation requires openpyxl library")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = job.target.value.title()

        # Style for headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Handle timeline dict format
        if isinstance(data, dict) and "events" in data:
            data = data.get("events", [])

        if not isinstance(data, list):
            data = [data]

        if not data:
            ws["A1"] = "No data to export"
            wb.save(file_path)
            return

        # Flatten data for spreadsheet
        flat_data = self._flatten_for_csv(data)

        # Get all unique keys
        all_keys = set()
        for record in flat_data:
            all_keys.update(record.keys())
        fieldnames = sorted(all_keys)

        # Write headers
        for col, header in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for row_idx, record in enumerate(flat_data, 2):
            for col_idx, header in enumerate(fieldnames, 1):
                value = record.get(header, "")
                # Truncate very long strings
                if isinstance(value, str) and len(value) > 32000:
                    value = value[:32000] + "..."
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths (approximate)
        for col_idx, header in enumerate(fieldnames, 1):
            max_length = len(header)
            for row in range(2, min(len(flat_data) + 2, 100)):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, min(len(str(cell_value)), 50))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

        wb.save(file_path)

    def _get_file_extension(self, format: ExportFormat) -> str:
        """Get file extension for format."""
        extensions = {
            ExportFormat.JSON: "json",
            ExportFormat.CSV: "csv",
            ExportFormat.PDF: "pdf",
            ExportFormat.DOCX: "docx",
            ExportFormat.XLSX: "xlsx",
        }
        return extensions.get(format, "bin")

    async def _save_job(self, job: ExportJob, update: bool = False) -> None:
        """Save export job to database."""
        if not self._db:
            return

        # Get tenant_id for multi-tenancy
        tenant_id = self.get_tenant_id_or_none()

        # Serialize options
        options_json = "{}"
        if job.options:
            options_json = json.dumps({
                "include_metadata": job.options.include_metadata,
                "include_relationships": job.options.include_relationships,
                "date_range_start": job.options.date_range_start.isoformat() if job.options.date_range_start else None,
                "date_range_end": job.options.date_range_end.isoformat() if job.options.date_range_end else None,
                "entity_types": job.options.entity_types,
                "flatten": job.options.flatten,
                "max_records": job.options.max_records,
                "sort_by": job.options.sort_by,
                "sort_order": job.options.sort_order,
                # Timeline-specific options
                "include_conflicts": job.options.include_conflicts,
                "include_gaps": job.options.include_gaps,
                "group_by": job.options.group_by,
            })

        data = (
            job.id,
            job.format.value,
            job.target.value,
            job.status.value,
            job.created_at.isoformat(),
            job.started_at.isoformat() if job.started_at else None,
            job.completed_at.isoformat() if job.completed_at else None,
            job.file_path,
            job.file_size,
            job.download_url,
            job.expires_at.isoformat() if job.expires_at else None,
            job.error,
            json.dumps(job.filters),
            options_json,
            job.record_count,
            job.processing_time_ms,
            job.created_by,
            json.dumps(job.metadata),
            str(tenant_id) if tenant_id else None,
        )

        if update:
            # For updates, also filter by tenant_id
            if tenant_id:
                await self._db.execute("""
                    UPDATE arkham_export_jobs SET
                        format=?, target=?, status=?, created_at=?, started_at=?,
                        completed_at=?, file_path=?, file_size=?, download_url=?,
                        expires_at=?, error=?, filters=?, options=?,
                        record_count=?, processing_time_ms=?, created_by=?, metadata=?,
                        tenant_id=?
                    WHERE id=? AND tenant_id=?
                """, data[1:] + (job.id, str(tenant_id)))
            else:
                await self._db.execute("""
                    UPDATE arkham_export_jobs SET
                        format=?, target=?, status=?, created_at=?, started_at=?,
                        completed_at=?, file_path=?, file_size=?, download_url=?,
                        expires_at=?, error=?, filters=?, options=?,
                        record_count=?, processing_time_ms=?, created_by=?, metadata=?,
                        tenant_id=?
                    WHERE id=?
                """, data[1:] + (job.id,))
        else:
            await self._db.execute("""
                INSERT INTO arkham_export_jobs (
                    id, format, target, status, created_at, started_at,
                    completed_at, file_path, file_size, download_url,
                    expires_at, error, filters, options,
                    record_count, processing_time_ms, created_by, metadata,
                    tenant_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, data)

    def _row_to_job(self, row: Dict[str, Any]) -> ExportJob:
        """Convert database row to ExportJob object."""
        # Parse options
        options = None
        if row.get("options"):
            options_data = json.loads(row["options"])
            options = ExportOptions(
                include_metadata=options_data.get("include_metadata", True),
                include_relationships=options_data.get("include_relationships", True),
                date_range_start=datetime.fromisoformat(options_data["date_range_start"]) if options_data.get("date_range_start") else None,
                date_range_end=datetime.fromisoformat(options_data["date_range_end"]) if options_data.get("date_range_end") else None,
                entity_types=options_data.get("entity_types"),
                flatten=options_data.get("flatten", False),
                max_records=options_data.get("max_records"),
                sort_by=options_data.get("sort_by"),
                sort_order=options_data.get("sort_order", "asc"),
                # Timeline-specific options
                include_conflicts=options_data.get("include_conflicts", False),
                include_gaps=options_data.get("include_gaps", False),
                group_by=options_data.get("group_by"),
            )

        return ExportJob(
            id=row["id"],
            format=ExportFormat(row["format"]),
            target=ExportTarget(row["target"]),
            status=ExportStatus(row["status"]),
            created_at=datetime.fromisoformat(row["created_at"]) if row["created_at"] else datetime.utcnow(),
            started_at=datetime.fromisoformat(row["started_at"]) if row["started_at"] else None,
            completed_at=datetime.fromisoformat(row["completed_at"]) if row["completed_at"] else None,
            file_path=row["file_path"],
            file_size=row["file_size"],
            download_url=row["download_url"],
            expires_at=datetime.fromisoformat(row["expires_at"]) if row["expires_at"] else None,
            error=row["error"],
            filters=json.loads(row["filters"] or "{}"),
            options=options,
            record_count=row["record_count"],
            processing_time_ms=row["processing_time_ms"],
            created_by=row["created_by"],
            metadata=json.loads(row["metadata"] or "{}"),
        )
